import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from db_config import create_db_engine

# --- CONFIGURAÇÕES ---
NOME_ARQUIVO_XLSX = "indice_analitico.xlsx"
NOME_ARQUIVO_XLSX_COMPARATIVO = "indice_comparativo.xlsx"


def extrair_indice_do_banco():
    """Conecta ao banco e extrai orgao, disciplina, assunto e arquivo_fonte."""
    print("Conectando ao banco de dados...")
    try:
        engine = create_db_engine()
        df = pd.read_sql_query(
            "SELECT orgao, disciplina, assunto, arquivo_fonte FROM informativos", engine
        )
        print("Dados extraídos com sucesso!")
        return df
    except Exception as e:
        print(f"\n!!!! ERRO AO CONECTAR OU EXTRAIR DADOS !!!!")
        print(f"Detalhes do erro: {e}")
        return None


def montar_disciplinas_dict(df):
    """
    Retorna {disciplina: {'STF': {assunto: {nums}}, 'STJ': {assunto: {nums}}}}
    """
    disciplinas_dict = {}
    df = df.copy()
    df["num_inf"] = pd.to_numeric(
        df["arquivo_fonte"].str.extract(r"(\d+)")[0], errors="coerce"
    ).astype("Int64")

    for _, row in df.iterrows():
        disciplina = row["disciplina"]
        orgao = row["orgao"]
        assunto = row["assunto"]
        num_inf = row["num_inf"]

        if pd.isna(num_inf) or orgao not in ("STF", "STJ"):
            continue

        if disciplina not in disciplinas_dict:
            disciplinas_dict[disciplina] = {"STF": {}, "STJ": {}}

        nums = disciplinas_dict[disciplina][orgao].setdefault(assunto, set())
        nums.add(int(num_inf))

    return disciplinas_dict


def _formatar_assunto_com_informativos(assunto, numeros):
    nums = sorted(numeros, key=int)
    return f"{assunto} ({', '.join(map(str, nums))})"


def _aplicar_estilos_planilha(ws, titulo_aba):
    ws.title = titulo_aba
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    disciplina_font = Font(bold=True, color="4F81BD")

    ws["A1"] = "STF"
    ws["B1"] = "STJ"
    for cell in (ws["A1"], ws["B1"]):
        cell.font = header_font
        cell.fill = header_fill

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80
    return disciplina_font


def _preencher_planilha(ws, disciplinas_dict, com_informativos):
    disciplina_font = _aplicar_estilos_planilha(
        ws, "Índice com Informativos" if com_informativos else "Índice Comparativo"
    )
    current_row = 2

    for disciplina in sorted(disciplinas_dict.keys()):
        assuntos_stf = disciplinas_dict[disciplina]["STF"]
        assuntos_stj = disciplinas_dict[disciplina]["STJ"]

        if com_informativos:
            temas_stf = [
                _formatar_assunto_com_informativos(a, assuntos_stf[a])
                for a in sorted(assuntos_stf.keys())
            ]
            temas_stj = [
                _formatar_assunto_com_informativos(a, assuntos_stj[a])
                for a in sorted(assuntos_stj.keys())
            ]
        else:
            temas_stf = sorted(assuntos_stf.keys())
            temas_stj = sorted(assuntos_stj.keys())

        if not temas_stf and not temas_stj:
            continue

        cell_stf = ws.cell(row=current_row, column=1, value=disciplina.upper())
        cell_stf.font = disciplina_font
        cell_stj = ws.cell(row=current_row, column=2, value=disciplina.upper())
        cell_stj.font = disciplina_font
        current_row += 1

        max_len = max(len(temas_stf), len(temas_stj))
        for i in range(max_len):
            if i < len(temas_stf):
                ws.cell(row=current_row + i, column=1, value=temas_stf[i])
            if i < len(temas_stj):
                ws.cell(row=current_row + i, column=2, value=temas_stj[i])

        current_row += max_len + 1


def gerar_arquivo_xlsx_com_informativos(disciplinas_dict):
    """Gera Excel com assuntos e números dos informativos por órgão."""
    print(f"Gerando o arquivo '{NOME_ARQUIVO_XLSX}'...")
    wb = Workbook()
    _preencher_planilha(wb.active, disciplinas_dict, com_informativos=True)
    wb.save(NOME_ARQUIVO_XLSX)
    print(f"Arquivo '{NOME_ARQUIVO_XLSX}' criado com sucesso.")


def gerar_arquivo_xlsx_comparativo(disciplinas_dict):
    """Gera Excel comparativo STF x STJ apenas com nomes dos assuntos."""
    print(f"Gerando o arquivo '{NOME_ARQUIVO_XLSX_COMPARATIVO}'...")
    wb = Workbook()
    _preencher_planilha(wb.active, disciplinas_dict, com_informativos=False)
    wb.save(NOME_ARQUIVO_XLSX_COMPARATIVO)
    print(f"Arquivo '{NOME_ARQUIVO_XLSX_COMPARATIVO}' criado com sucesso.")


if __name__ == "__main__":
    dataframe_informativos = extrair_indice_do_banco()
    if dataframe_informativos is not None:
        dataframe_informativos.dropna(
            subset=["orgao", "disciplina", "assunto", "arquivo_fonte"], inplace=True
        )
        indice = montar_disciplinas_dict(dataframe_informativos)
        gerar_arquivo_xlsx_com_informativos(indice)
        print("-" * 20)
        gerar_arquivo_xlsx_comparativo(indice)
