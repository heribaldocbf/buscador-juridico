import pandas as pd
from pathlib import Path
from docx import Document
from openpyxl import load_workbook

# --- CONFIGURAÇÕES ---
# 1. Defina a pasta onde estão os NOVOS informativos (.docx) que você quer analisar.
PASTA_NOVOS_INFORMATIVOS = r"C:\Users\Heribaldo\Desktop\BD"

# 2. Defina o nome do seu arquivo Excel que serve como índice mestre.
ARQUIVO_INDICE_XLSX = "indice_analitico.xlsx"

# 3. Nome do arquivo de texto que será gerado com a análise.
ARQUIVO_RELATORIO_SAIDA = "relatorio_correspondencia.txt"

def ler_indice_mestre_xlsx(caminho_arquivo):
    """
    Lê o arquivo Excel e cria um dicionário de referência com todos os
    assuntos existentes para cada disciplina e órgão.
    Retorna: {disciplina: {'STF': {assuntos}, 'STJ': {assuntos}}}
    """
    print(f"Lendo o índice mestre: '{caminho_arquivo}'...")
    if not Path(caminho_arquivo).exists():
        print(f"ERRO: Arquivo de índice '{caminho_arquivo}' não encontrado.")
        return None

    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        indice_mestre = {}
        disciplina_atual_stf = None
        disciplina_atual_stj = None

        for row in ws.iter_rows(min_row=2): # Pula o cabeçalho
            valor_stf = row[0].value
            valor_stj = row[1].value
            
            # Checa coluna do STF
            if valor_stf:
                if row[0].font and row[0].font.bold:
                    disciplina_atual_stf = valor_stf.strip().upper()
                    if disciplina_atual_stf not in indice_mestre:
                        indice_mestre[disciplina_atual_stf] = {'STF': set(), 'STJ': set()}
                elif disciplina_atual_stf:
                    indice_mestre[disciplina_atual_stf]['STF'].add(valor_stf.strip())

            # Checa coluna do STJ
            if valor_stj:
                if row[1].font and row[1].font.bold:
                    disciplina_atual_stj = valor_stj.strip().upper()
                    if disciplina_atual_stj not in indice_mestre:
                        indice_mestre[disciplina_atual_stj] = {'STF': set(), 'STJ': set()}
                elif disciplina_atual_stj:
                    indice_mestre[disciplina_atual_stj]['STJ'].add(valor_stj.strip())
        
        print("Leitura do índice mestre concluída.")
        return indice_mestre
    except Exception as e:
        print(f"ERRO ao ler o arquivo Excel: {e}")
        return None


def processar_novos_informativos(caminho_pasta):
    """
    Lê os arquivos .docx de uma pasta e extrai as informações de
    disciplina, assunto, órgão e nome do arquivo.
    """
    print(f"Processando novos informativos da pasta: '{caminho_pasta}'...")
    p = Path(caminho_pasta)
    if not p.exists():
        print(f"ERRO: A pasta de novos informativos não foi encontrada.")
        return []

    dados_extraidos = []
    lista_de_docx = list(p.rglob("*.docx"))
    print(f"Encontrados {len(lista_de_docx)} arquivos DOCX para processar.")

    for docx_path in lista_de_docx:
        try:
            nome_arquivo, orgao = docx_path.name, docx_path.parent.name.upper()
            if orgao not in ['STF', 'STJ']:
                orgao = 'STF' if 'stf' in nome_arquivo.lower() else 'STJ' if 'stj' in nome_arquivo.lower() else 'DESCONHECIDO'

            document = Document(docx_path)
            current_disciplina = None
            capturando_tese = False

            for para in document.paragraphs:
                texto_paragrafo = para.text.strip()
                if not texto_paragrafo: continue

                estilo = para.style.name
                
                if estilo.startswith(('Heading 1', 'Título 1')):
                    current_disciplina = texto_paragrafo.upper()
                elif estilo.startswith(('Heading 2', 'Título 2')):
                    if current_disciplina:
                        dados_extraidos.append({
                            "arquivo_fonte": nome_arquivo,
                            "orgao": orgao,
                            "disciplina": current_disciplina,
                            "assunto": texto_paragrafo,
                        })
        except Exception as e:
            print(f"!!!! Erro ao processar o arquivo {docx_path.name}: {e}")
    
    print("Processamento de novos informativos concluído.")
    return dados_extraidos


def analisar_e_gerar_relatorio(indice_mestre, novos_dados):
    """
    Compara os novos dados com o índice mestre e gera um relatório
    de assuntos não correspondentes.
    """
    if not indice_mestre or not novos_dados:
        print("Análise não realizada por falta de dados (índice mestre ou novos informativos).")
        return

    print("Iniciando análise comparativa...")
    nao_correspondentes_stf = []
    nao_correspondentes_stj = []

    for item in novos_dados:
        disciplina = item['disciplina']
        assunto = item['assunto']
        orgao = item['orgao']
        
        # Verifica se o assunto está no índice mestre para a disciplina e órgão corretos
        if disciplina in indice_mestre and assunto not in indice_mestre[disciplina][orgao]:
            linha_relatorio = f"- Disciplina: {disciplina}\n  - Assunto Novo: '{assunto}'\n  - Fonte: {item['arquivo_fonte']}\n"
            if orgao == 'STF':
                nao_correspondentes_stf.append(linha_relatorio)
            elif orgao == 'STJ':
                nao_correspondentes_stj.append(linha_relatorio)

    print("Análise concluída. Gerando relatório...")
    with open(ARQUIVO_RELATORIO_SAIDA, 'w', encoding='utf-8') as f:
        f.write("RELATÓRIO DE CORRESPONDÊNCIA DE ASSUNTOS\n")
        f.write("="*40 + "\n\n")

        f.write("--- ASSUNTOS NOVOS OU NÃO CORRESPONDENTES NO STF ---\n")
        if nao_correspondentes_stf:
            for entrada in sorted(nao_correspondentes_stf):
                f.write(entrada)
        else:
            f.write("Nenhum assunto novo encontrado para o STF.\n")

        f.write("\n\n--- ASSUNTOS NOVOS OU NÃO CORRESPONDENTES NO STJ ---\n")
        if nao_correspondentes_stj:
            for entrada in sorted(nao_correspondentes_stj):
                f.write(entrada)
        else:
            f.write("Nenhum assunto novo encontrado para o STJ.\n")
            
    print(f"Relatório salvo em '{ARQUIVO_RELATORIO_SAIDA}'.")


if __name__ == "__main__":
    indice_referencia = ler_indice_mestre_xlsx(ARQUIVO_INDICE_XLSX)
    dados_novos = processar_novos_informativos(PASTA_NOVOS_INFORMATIVOS)
    analisar_e_gerar_relatorio(indice_referencia, dados_novos)